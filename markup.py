#!/usr/bin/env python3
"""
inject_excel_to_bc_report.py

Reads an Excel file (skipping first 3 rows) with columns: ID, S, Notes
and injects S and Notes values into a Beyond Compare HTML report,
matching rows by <td id="markupID_N"> elements.

Handles merged cells: if multiple IDs share the same S / Notes value,
the first occurrence gets rowspan=N and subsequent cells are omitted,
matching the merged appearance.

Usage:
    python inject_excel_to_bc_report.py data.xlsx BcReport.html

    Output is saved next to the HTML file as BcReport_annotated.html

    # Optional: specify which sheet (default: first sheet)
    python inject_excel_to_bc_report.py data.xlsx BcReport.html --sheet "Sheet1"

Requirements:
    pip install openpyxl beautifulsoup4 lxml
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")

try:
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("Missing dependency: pip install beautifulsoup4 lxml")


# ---------------------------------------------------------------------------
# 1. Read Excel
# ---------------------------------------------------------------------------

def read_excel(path: str, sheet_name=None) -> dict:
    """
    Read the Excel file, skipping the first 3 rows.
    Expects columns in order: ID, S, Notes  (columns A, B, C).
    Handles merged cells by forward-filling their values.

    Returns a dict: { id_value (int): {"s": ..., "notes": ...} }
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # --- Resolve merged cell ranges so every cell has a value ---
    # openpyxl leaves merged (non-anchor) cells as None; we fill them
    # with the anchor cell's value.
    merged_map = {}  # (row, col) -> (anchor_row, anchor_col)
    for merge_range in ws.merged_cells.ranges:
        anchor_row = merge_range.min_row
        anchor_col = merge_range.min_col
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                if (row, col) != (anchor_row, anchor_col):
                    merged_map[(row, col)] = (anchor_row, anchor_col)

    def cell_value(ws, row, col):
        coord = (row, col)
        if coord in merged_map:
            ar, ac = merged_map[coord]
            return ws.cell(ar, ac).value
        return ws.cell(row, col).value

    # Rows 1-3 are skipped; data starts at row 4
    DATA_START_ROW = 4  # 1-indexed in openpyxl

    records = {}
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        id_val   = cell_value(ws, row_idx, 1)   # Column A = ID
        s_val    = cell_value(ws, row_idx, 2)   # Column B = S
        notes_val = cell_value(ws, row_idx, 3)  # Column C = Notes

        # Stop at first completely empty row
        if id_val is None and s_val is None and notes_val is None:
            break

        if id_val is None:
            continue

        try:
            id_int = int(id_val)
        except (ValueError, TypeError):
            print(f"  [WARN] Row {row_idx}: ID '{id_val}' is not an integer — skipped.")
            continue

        records[id_int] = {
            "s":     str(s_val)     if s_val     is not None else "",
            "notes": str(notes_val) if notes_val is not None else "",
        }

    print(f"  Loaded {len(records)} records from Excel (IDs {min(records)} – {max(records)}).")
    return records


# ---------------------------------------------------------------------------
# 2. Parse HTML and find markupID rows
# ---------------------------------------------------------------------------

def find_markup_rows(soup: BeautifulSoup):
    """
    Find all <td> elements with id matching the exact pattern markupID_N
    where N is a positive integer (e.g. markupID_1, markupID_42).

    Returns a list of (id_number, td_element) tuples, sorted by id_number.
    """
    import re
    pattern = re.compile(r'^markupID_(\d+)$')
    results = []
    for td in soup.find_all("td", id=True):
        m = pattern.match(td["id"])
        if m:
            results.append((int(m.group(1)), td))
        
    results.sort(key=lambda x: x[0])
    return results


# ---------------------------------------------------------------------------
# 3. Compute merged-cell groups
# ---------------------------------------------------------------------------

def compute_groups(records: dict, markup_ids: list) -> dict:
    """
    For each markupID present in the HTML, determine if consecutive IDs
    share the same (s, notes) value — indicating merged cells in Excel.

    Returns a dict:
        { id_number: {"s": ..., "notes": ..., "rowspan": N, "render": bool} }

    render=False means this row's S/Notes cells should be omitted
    (they were covered by a rowspan from an earlier row).
    """
    present_ids = sorted([n for n, _ in markup_ids if n in records])

    # Group consecutive IDs with identical (s, notes)
    groups = []
    i = 0
    while i < len(present_ids):
        current_id = present_ids[i]
        current_rec = records[current_id]
        span = 1
        j = i + 1
        while j < len(present_ids):
            next_id = present_ids[j]
            # Only group if IDs are truly consecutive
            if next_id == present_ids[j - 1] + 1 and records[next_id] == current_rec:
                span += 1
                j += 1
            else:
                break
        for k in range(span):
            gid = present_ids[i + k]
            groups.append({
                "id": gid,
                "s": current_rec["s"],
                "notes": current_rec["notes"],
                "rowspan": span if k == 0 else 0,  # 0 = suppressed
                "render": k == 0,
            })
        i = j

    return {g["id"]: g for g in groups}


# ---------------------------------------------------------------------------
# 4. Inject into HTML
# ---------------------------------------------------------------------------

STYLE_BLOCK = """
<style>
/* Injected by inject_excel_to_bc_report.py */
.injected-s {
    background-color: #e8f4fd;
    border: 1px solid #b0d4f1;
    padding: 2px 6px;
    font-size: 0.9em;
    vertical-align: middle;
    text-align: center;
    white-space: pre-wrap;
    word-break: break-word;
    max-width: 120px;
}
.injected-notes {
    background-color: #fefce8;
    border: 1px solid #e2d96e;
    padding: 2px 6px;
    font-size: 0.9em;
    vertical-align: middle;
    white-space: pre-wrap;
    word-break: break-word;
    max-width: 200px;
}
.injected-empty {
    background-color: #f9f9f9;
    border: 1px solid #ddd;
    padding: 2px 6px;
    color: #aaa;
    font-size: 0.85em;
    vertical-align: middle;
    text-align: center;
}
</style>
"""

def inject(soup: BeautifulSoup, markup_rows: list, group_map: dict):
    """
    For each markupID <td>, insert S and Notes <td> elements
    immediately after it in the same <tr>.
    Applies rowspan for merged groups.
    """
    injected = 0
    skipped  = 0
    missing  = 0

    for n, td in markup_rows:
        tr = td.parent
        if tr is None or tr.name != "tr":
            print(f"  [WARN] markupID_{n}: parent is not a <tr> — skipped.")
            continue

        if n not in group_map:
            # ID exists in HTML but not in Excel — insert placeholder
            s_td = soup.new_tag("td", attrs={"class": "injected-empty"})
            s_td.string = "—"
            n_td = soup.new_tag("td", attrs={"class": "injected-empty"})
            n_td.string = "—"
            td.insert_after(n_td)
            td.insert_after(s_td)
            missing += 1
            continue

        g = group_map[n]

        if not g["render"]:
            # This row is covered by a rowspan above — don't add cells
            skipped += 1
            continue

        # Build S cell
        s_td = soup.new_tag("td", attrs={"class": "injected-s"})
        if g["rowspan"] > 1:
            s_td["rowspan"] = str(g["rowspan"])
        s_td.string = g["s"]

        # Build Notes cell
        n_td = soup.new_tag("td", attrs={"class": "injected-notes"})
        if g["rowspan"] > 1:
            n_td["rowspan"] = str(g["rowspan"])
        n_td.string = g["notes"]

        # Insert after the markupID td (order: td → s_td → n_td)
        td.insert_after(n_td)
        td.insert_after(s_td)
        injected += 1

    print(f"  Injected: {injected} rows | Rowspan-skipped: {skipped} | No Excel data: {missing}")


def add_header_columns(soup: BeautifulSoup):
    """
    Add 'S' and 'Notes' header columns to any header row that has a
    cell containing 'markupID' (Beyond Compare header rows sometimes
    differ — we do a best-effort search).
    """
    # Look for any <th> or header <td> that mentions the file names,
    # since BC doesn't always use a conventional <thead>.
    # We add headers to the FIRST <tr> in the table.
    tables = soup.find_all("table", class_="fc")
    for table in tables:
        first_tr = table.find("tr")
        if first_tr:
            s_th = soup.new_tag("th")
            s_th.string = "S"
            n_th = soup.new_tag("th")
            n_th.string = "Notes"
            first_tr.append(s_th)
            first_tr.append(n_th)


# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("excel", help="Path to the .xlsx file")
    parser.add_argument("html",  help="Path to the Beyond Compare HTML report")
    parser.add_argument("--sheet", default=None, help="Excel sheet name (default: first sheet)")
    parser.add_argument("--no-header", action="store_true",
                        help="Skip adding S/Notes header columns")
    args = parser.parse_args()

    # Validate inputs
    excel_path = Path(args.excel)
    html_path  = Path(args.html)

    if not excel_path.exists():
        sys.exit(f"Excel file not found: {excel_path}")
    if not html_path.exists():
        sys.exit(f"HTML file not found: {html_path}")

    # Auto-generate output path: same directory as HTML, with _annotated suffix
    out_path = html_path.with_stem(html_path.stem + "_annotated")

    print(f"\n[1/5] Reading Excel: {excel_path}")
    records = read_excel(str(excel_path), sheet_name=args.sheet)
    if not records:
        sys.exit("No records found in Excel. Check that data starts at row 4 (after 3 header rows).")

    print(f"\n[2/5] Parsing HTML: {html_path}")
    html_text = html_path.read_text(encoding="utf-8", errors="replace")
    soup = BeautifulSoup(html_text, "lxml")

    print(f"\n[3/5] Locating markupID_N elements...")
    markup_rows = find_markup_rows(soup)
    if not markup_rows:
        sys.exit("No <td id='markupID_N'> elements found in the HTML.\n"
                 "Hint: open the report in a browser, inspect a data row, and confirm the id pattern.")
    print(f"  Found {len(markup_rows)} markupID elements "
          f"(IDs {markup_rows[0][0]}–{markup_rows[-1][0]}).")

    print(f"\n[4/5] Computing merged-cell groups...")
    group_map = compute_groups(records, markup_rows)
    n_merged_groups = sum(1 for g in group_map.values() if g["rowspan"] > 1)
    print(f"  {len(group_map)} IDs mapped | {n_merged_groups} merged groups.")

    print(f"\n[5/5] Injecting columns into HTML...")
    head = soup.find("head")
    if head:
        head.append(BeautifulSoup(STYLE_BLOCK, "lxml").find("style"))
    else:
        soup.find("body").insert(0, BeautifulSoup(STYLE_BLOCK, "lxml").find("style"))

    if not args.no_header:
        add_header_columns(soup)

    inject(soup, markup_rows, group_map)

    out_path.write_text(str(soup), encoding="utf-8")
    print(f"\n✓ Annotated report saved to: {out_path}\n")


if __name__ == "__main__":
    main()
